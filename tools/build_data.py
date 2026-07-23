#!/usr/bin/env python3
"""
build_data.py — convert the content spreadsheets into the app's JSON data files.

Usage:
    pip install openpyxl        # one-time
    python3 tools/build_data.py # run from the repo root (paths are resolved automatically)

Reads every workbook in  content/*.xlsx  and writes:
    data/books.json                 (catalog of all books, ordered)
    data/<book-id>/book.json        (book meta + chapter list)
    data/<book-id>/ch<n>.json       (per-chapter content)

The spreadsheet is the SOURCE OF TRUTH; data/** is generated — do not hand-edit it.
All input is validated first; if anything is wrong, the script prints plain-language
errors and writes NOTHING. See CLAUDE.md §4 for the sheet/column layout.
"""
import base64
import hashlib
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Missing dependency. Run:  pip install openpyxl")

try:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.primitives import hashes
except ImportError:
    sys.exit("Missing dependency. Run:  pip install cryptography")

# --- content encryption (matches the browser's Web Crypto in gate.js / app.js) ---
# Each book's content is encrypted with AES-256-GCM using a key derived from the
# book's access code via PBKDF2-HMAC-SHA256. The site only ever serves ciphertext;
# the code never leaves the spreadsheet. Salt/IV are derived deterministically so
# unchanged content produces identical output (clean git diffs).
PBKDF2_ITERS = 200000

def book_salt(bid):
    return hashlib.sha256(("3r-salt:" + bid).encode()).digest()[:16]

def derive_key(code, salt):
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=PBKDF2_ITERS)
    return kdf.derive(norm_code(code).encode("utf-8"))

def encrypt_obj(obj, key, ivseed):
    pt = json.dumps(obj, ensure_ascii=False).encode("utf-8")
    iv = hashlib.sha256(("3r-iv:" + ivseed).encode("utf-8") + pt).digest()[:12]
    ct = AESGCM(key).encrypt(iv, pt, None)  # ciphertext includes the 16-byte GCM tag
    return {"v": 1, "iv": base64.b64encode(iv).decode(), "ct": base64.b64encode(ct).decode()}

ROOT = Path(__file__).resolve().parent.parent      # repo root
CONTENT = ROOT / "content"                          # source spreadsheets (not web-served)
DATA = ROOT / "docs" / "data"                       # generated JSON, inside the published site

errors = []
def err(msg): errors.append(msg)

def norm_code(s):
    return (s or "").upper().replace(" ", "").replace("-", "")

def as_bool(v):
    if isinstance(v, bool): return v
    return str(v).strip().lower() in ("true", "1", "yes", "y")

def as_int(v, where):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        err("%s: expected a number but got %r" % (where, v))
        return None

def sheet_rows(wb, name, where):
    """Return a list of dicts keyed by the header row. Missing sheet -> []."""
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue  # skip blank rows
        d = {}
        for i, h in enumerate(headers):
            if h:
                d[h] = r[i] if i < len(r) else None
        out.append(d)
    return out

def cell(d, key):
    v = d.get(key)
    if v is None: return ""
    return str(v).strip()

def build_content(nrows, crows, mrows, vrows, where, label):
    """Build one study unit's {videos, notes, decks, mcqs} from its filtered rows.
    Used for a whole chapter (no subtopics) or for a single subtopic. `label` is a
    human string like 'ch1' or 'ch1/central-tendency' used only in error messages."""
    videos = [{"title": cell(v, "title"), "yt": cell(v, "yt")} for v in vrows]
    notes = [cell(x, "note") for x in nrows if cell(x, "note")]
    # decks (preserve first-seen order)
    decks, order_ids = {}, []
    for row in crows:
        did = cell(row, "deck_id")
        front, back = cell(row, "front"), cell(row, "back")
        if not did:
            err("%s %s: a flashcard row has no deck_id" % (where, label)); continue
        if not front or not back:
            err("%s %s deck '%s': a card is missing front or back" % (where, label, did)); continue
        if did not in decks:
            decks[did] = {"id": did, "name": cell(row, "deck_name"),
                          "desc": cell(row, "deck_desc"), "cards": []}
            order_ids.append(did)
        decks[did]["cards"].append([front, back])
    deck_list = [decks[d] for d in order_ids]
    # mcqs
    mcqs = []
    for row in mrows:
        q = cell(row, "question")
        ci = as_int(row.get("correct"), "%s %s MCQ" % (where, label))
        if not q or ci is None:
            continue
        if ci < 1 or ci > 4:
            err("%s %s: MCQ '%s' has correct=%s (must be 1-4)" % (where, label, q[:40], ci)); continue
        opts = []
        for i in (1, 2, 3, 4):
            text = cell(row, "option%d" % i)
            why = cell(row, "why%d" % i)
            if not text:
                continue
            if not why:
                err("%s %s: MCQ '%s' option%d has no explanation (why%d)" % (where, label, q[:40], i, i))
            is_correct = (i == ci)
            # the correct option's explanation is auto-prefixed with "Correct —"
            if is_correct and why and not why.lower().startswith("correct"):
                why = "Correct — " + why
            opts.append({"text": text, "correct": is_correct, "why": why})
        if not any(o["correct"] for o in opts):
            err("%s %s: MCQ '%s' — 'correct' points to an empty option" % (where, label, q[:40]))
        mcqs.append({"q": q, "options": opts})
    return {"videos": videos, "notes": notes, "decks": deck_list, "mcqs": mcqs}

def build_book(path):
    """Parse one workbook into (catalog_entry, book_json, {chapnum: chapter_json})."""
    where = path.name
    wb = load_workbook(path, data_only=True, read_only=True)

    book_rows = sheet_rows(wb, "Book", where)
    if not book_rows:
        err("%s: missing or empty 'Book' sheet" % where)
        return None
    b = book_rows[0]
    bid = cell(b, "id")
    if not bid:
        err("%s: Book.id is required" % where)
        return None
    ready = as_bool(b.get("ready"))
    code = norm_code(cell(b, "code"))
    if ready and not code:
        err("%s (%s): a ready book must have an access code" % (where, bid))

    catalog = {
        "id": bid,
        "name": cell(b, "name"),
        "board": cell(b, "board"),
        "desc": cell(b, "desc"),
        "category": (cell(b, "category") or "other").lower(),
        "ready": ready,
    }
    # A ready, code-protected book publishes its per-book salt (NOT the code) so the
    # browser can derive the decryption key once the student enters the code.
    salt = book_salt(bid) if (ready and code) else None
    if salt:
        catalog["salt"] = base64.b64encode(salt).decode()
        catalog["locked"] = True
    order = as_int(b.get("order"), "%s: Book.order" % where) if b.get("order") not in (None, "") else 999

    # chapters
    chapters = []
    for c in sheet_rows(wb, "Chapters", where):
        n = as_int(c.get("n"), "%s: Chapters.n" % where)
        if n is None:
            continue
        chapters.append({"n": n, "title": cell(c, "title"), "ready": as_bool(c.get("ready")),
                         "kind": (cell(c, "kind") or "chapter").lower()})
    chapters.sort(key=lambda c: c["n"])

    # subtopics (optional second level within a chapter) -> {chapter n: [ordered subtopics]}
    subs_by = {}
    for s in sheet_rows(wb, "Subtopics", where):
        cn = as_int(s.get("chapter"), "%s: Subtopics.chapter" % where)
        if cn is None:
            continue
        sid = cell(s, "id")
        if not sid:
            err("%s: a Subtopics row (chapter %s) has no id" % (where, cn)); continue
        subs_by.setdefault(cn, []).append({
            "id": sid, "title": cell(s, "title"), "desc": cell(s, "desc"),
            "ready": as_bool(s.get("ready")),
            "order": as_int(s.get("order"), "%s: Subtopics.order" % where) if s.get("order") not in (None, "") else 999,
        })
    for cn, slist in subs_by.items():
        slist.sort(key=lambda s: (s["order"], s["id"]))
        seen = set()
        for s in slist:
            if s["id"] in seen:
                err("%s ch%s: duplicate subtopic id '%s'" % (where, cn, s["id"]))
            seen.add(s["id"])
    # expose subtopic titles on the public TOC (book.json) and validate readiness
    for c in chapters:
        slist = subs_by.get(c["n"])
        if slist:
            c["subtopics"] = [{"id": s["id"], "title": s["title"], "ready": s["ready"]} for s in slist]
            if c["ready"] and not any(s["ready"] for s in slist):
                err("%s ch%s: chapter is ready but has no ready subtopic" % (where, c["n"]))

    book_json = {
        "book": bid,
        "name": cell(b, "title") or cell(b, "name"),
        "edition": cell(b, "edition"),
        "tagline": cell(b, "tagline"),
        "chapters": chapters,
    }

    # group content rows by chapter number (then optionally by subtopic within a chapter)
    def group(sheet):
        g = {}
        for row in sheet_rows(wb, sheet, where):
            n = as_int(row.get("chapter"), "%s: %s.chapter" % (where, sheet))
            if n is None:
                continue
            g.setdefault(n, []).append(row)
        return g

    notes_by = group("Notes")
    cards_by = group("Flashcards")
    mcqs_by = group("MCQs")
    videos_by = group("Videos")

    def by_sub(rows, sid):
        return [r for r in rows if cell(r, "subtopic") == sid]

    chapter_json = {}
    chap_titles = {c["n"]: c["title"] for c in chapters}
    chap_kind = {c["n"]: c.get("kind", "chapter") for c in chapters}
    involved = set(notes_by) | set(cards_by) | set(mcqs_by) | set(videos_by) | set(subs_by)
    involved |= {c["n"] for c in chapters if c["ready"]}

    for n in sorted(involved):
        base = {"book": bid, "number": n, "title": chap_titles.get(n, "Chapter %d" % n),
                "kind": chap_kind.get(n, "chapter")}
        subs = subs_by.get(n)
        if subs:
            sub_ids = {s["id"] for s in subs}
            # every content row in a subtopic-chapter must name a declared subtopic
            for sheet_name, rows_by in (("Notes", notes_by), ("Flashcards", cards_by),
                                        ("MCQs", mcqs_by), ("Videos", videos_by)):
                for r in rows_by.get(n, []):
                    sid = cell(r, "subtopic")
                    if not sid:
                        err("%s ch%s: a %s row has no subtopic (this chapter uses subtopics)" % (where, n, sheet_name))
                    elif sid not in sub_ids:
                        err("%s ch%s: a %s row names unknown subtopic '%s'" % (where, n, sheet_name, sid))
            sub_list = []
            for s in subs:
                content = build_content(
                    by_sub(notes_by.get(n, []), s["id"]),
                    by_sub(cards_by.get(n, []), s["id"]),
                    by_sub(mcqs_by.get(n, []), s["id"]),
                    by_sub(videos_by.get(n, []), s["id"]),
                    where, "ch%s/%s" % (n, s["id"]))
                sub_list.append({"id": s["id"], "title": s["title"], "desc": s["desc"],
                                 "ready": s["ready"], **content})
            base["subtopics"] = sub_list
        else:
            base.update(build_content(notes_by.get(n, []), cards_by.get(n, []),
                                      mcqs_by.get(n, []), videos_by.get(n, []), where, "ch%s" % n))
        chapter_json[n] = base

    return {"catalog": catalog, "order": order, "book_json": book_json,
            "chapters": chapter_json, "code": code, "salt": salt}

def write_json(path, obj):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

def main():
    if not CONTENT.exists():
        sys.exit("No content/ folder found at %s" % CONTENT)
    books = []
    for xlsx in sorted(CONTENT.glob("*.xlsx")):
        if xlsx.name.startswith("~$"):   # skip Excel lock files
            continue
        # skip non-source copies (reviewer exports / backups) so they don't create
        # duplicate or code-less "books". Keep such copies in content/backups/ too.
        up = xlsx.stem.upper()
        if "-REVIEW" in up or "-BACKUP" in up or ".BACKUP" in xlsx.name.upper():
            continue
        parsed = build_book(xlsx)
        if parsed:
            books.append(parsed)

    if errors:
        print("Found %d problem(s) — nothing was written:\n" % len(errors))
        for e in errors:
            print("  •", e)
        sys.exit(1)

    books.sort(key=lambda b: (b["order"], b["catalog"]["name"]))

    # write everything
    write_json(DATA / "books.json", [b["catalog"] for b in books])
    total_ch = 0
    for b in books:
        bid = b["catalog"]["id"]
        # Clear this book's previously-generated chapter files so removing/unpublishing a
        # chapter doesn't leave stale output behind (book.json + unlock.enc are rewritten below).
        bookdir = DATA / bid
        if bookdir.exists():
            for f in list(bookdir.glob("ch*.enc")) + list(bookdir.glob("ch*.json")):
                f.unlink()
        # Coming-soon books (no chapters yet) only contribute a catalog entry.
        if b["book_json"]["chapters"]:
            # book.json (chapter list / titles) stays plaintext — it's the public table of contents.
            write_json(DATA / bid / "book.json", b["book_json"])
        code, salt = b["code"], b["salt"]
        if code and salt:
            key = derive_key(code, salt)
            # unlock marker — the gate decrypts this to validate an entered code
            write_json(DATA / bid / "unlock.enc", encrypt_obj({"ok": 1, "book": bid}, key, bid + ":unlock"))
            for n, chj in b["chapters"].items():
                write_json(DATA / bid / ("ch%d.enc" % n), encrypt_obj(chj, key, "%s:ch%d" % (bid, n)))
                total_ch += 1
        else:
            # no code → not encrypted (only happens for content without a protecting code)
            for n, chj in b["chapters"].items():
                write_json(DATA / bid / ("ch%d.json" % n), chj)
                total_ch += 1
    print("OK — wrote books.json (%d books) and %d encrypted chapter file(s)." % (len(books), total_ch))

if __name__ == "__main__":
    main()
